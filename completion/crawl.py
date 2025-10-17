"""
네이버 뉴스 크롤러 - 엑셀 입출력 통합 버전
실행: python crawl.py input.xlsx
"""
from __future__ import annotations
import os
import re
import time
import random
import logging
from pathlib import Path
from datetime import datetime, date, timezone, timedelta
from email.utils import parsedate_to_datetime
from typing import List, Dict, Optional, Any
import requests
import html
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 로깅 설정
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


# ========================================
# 1. 네이버 뉴스 크롤러
# ========================================
class NaverNewsCrawler:
    """네이버 Open API 기반 뉴스 크롤러"""
    
    BASE_URL = "https://openapi.naver.com/v1/search/news.json"

    def __init__(self, client_id: str = None, client_secret: str = None, timeout: float = 10.0):
        # .env 파일 로드
        try:
            from dotenv import load_dotenv, find_dotenv
            dotenv_path = find_dotenv(filename=".env", usecwd=True)
            if not dotenv_path:
                here = Path(__file__).resolve()
                for p in [here.parent, here.parent.parent]:
                    cand = p / ".env"
                    if cand.exists():
                        dotenv_path = str(cand)
                        break
            if dotenv_path:
                load_dotenv(dotenv_path=dotenv_path, override=True)
        except Exception:
            pass

        # API 키 설정
        env_id = (os.getenv("NAVER_CLIENT_ID") or "").strip()
        env_secret = (os.getenv("NAVER_CLIENT_SECRET") or "").strip()
        
        self.client_id = (client_id or env_id).strip()
        self.client_secret = (client_secret or env_secret).strip()
        
        if not self.client_id or not self.client_secret:
            raise ValueError("네이버 API 키가 필요합니다. .env 파일을 확인하세요.")
        
        self.session = requests.Session()
        self.timeout = timeout
        self.headers = {
            "X-Naver-Client-Id": self.client_id,
            "X-Naver-Client-Secret": self.client_secret,
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        }

    def search_news(self, query: str, start: int = 1, display: int = 10, sort: str = "date") -> Optional[Dict]:
        """네이버 뉴스 API 호출"""
        params = {"query": query, "start": start, "display": display, "sort": sort}
        max_retries = 5
        retry_count = 0
        base_backoff = 1.0

        while retry_count < max_retries:
            try:
                resp = self.session.get(self.BASE_URL, headers=self.headers, params=params, timeout=self.timeout)
                
                if resp.status_code == 200:
                    return resp.json()
                
                if resp.status_code == 429:
                    wait_time = min(60.0, (base_backoff * (2 ** retry_count)) + random.uniform(0, 1))
                    logger.warning(f"Rate limit - {retry_count+1}/{max_retries} 재시도 ({wait_time:.1f}초 대기)")
                    time.sleep(wait_time)
                    retry_count += 1
                    continue
                
                if 500 <= resp.status_code < 600:
                    wait_time = min(60.0, (base_backoff * (2 ** retry_count)) + random.uniform(0, 1))
                    logger.warning(f"서버 오류 {resp.status_code} - 재시도 중...")
                    time.sleep(wait_time)
                    retry_count += 1
                    continue
                
                logger.error(f"API 실패 - status={resp.status_code}")
                return None
                
            except requests.exceptions.RequestException as e:
                wait_time = min(30.0, (base_backoff * (2 ** retry_count)) + random.uniform(0, 1))
                logger.warning(f"API 예외: {e} - 재시도 중...")
                time.sleep(wait_time)
                retry_count += 1
        
        return None

    def search_news_multiple_pages(self, query: str, max_results: int = 100) -> List[Dict]:
        """여러 페이지 수집"""
        all_items = []
        page_size = 100
        current_start = 1
        hard_cap = 1000  # 네이버 API 제한

        while current_start <= hard_cap:
            remaining = max(0, max_results - len(all_items))
            if remaining == 0:
                break
            
            display_count = min(page_size, remaining, hard_cap - current_start + 1)
            
            resp = self.search_news(query=query, start=current_start, display=display_count, sort="date")
            if not resp or "items" not in resp:
                break
            
            items = resp["items"]
            if not items:
                break
            
            all_items.extend(items)
            logger.info(f"[{query}] {len(all_items)}건 수집 중...")
            
            if len(items) < display_count:
                break
            
            current_start += display_count
            time.sleep(random.uniform(0.1, 0.3))
        
        return all_items

    @staticmethod
    def _remove_html_tags(text: str) -> str:
        """HTML 태그 제거"""
        clean = re.sub(r"<[^>]+>", "", text or "")
        return html.unescape(clean)

    @staticmethod
    def _parse_pub_date(pub_date_str: str) -> str:
        """pubDate를 YYYYMMDD 형식으로 변환"""
        try:
            dt = parsedate_to_datetime(pub_date_str)
            return dt.strftime("%Y%m%d")
        except:
            return datetime.now().strftime("%Y%m%d")

    def format_news_data(self, items: List[Dict], keyword: str) -> List[Dict]:
        """API 응답을 표준 포맷으로 변환"""
        formatted = []
        for item in items:
            formatted.append({
                "title": self._remove_html_tags(item.get("title", "")),
                "description": self._remove_html_tags(item.get("description", "")),
                "link": item.get("link", ""),
                "original_link": item.get("originallink", ""),
                "pub_date": self._parse_pub_date(item.get("pubDate", "")),
                "keyword": keyword,
            })
        return formatted

    def close(self):
        try:
            self.session.close()
        except:
            pass


# ========================================
# 2. 엑셀 입출력 핸들러
# ========================================
class ExcelHandler:
    """엑셀 입출력 통합 처리"""

    @staticmethod
    def read_companies(filepath: str) -> Dict[str, Dict[str, str]]:
        """
        Company 시트에서 회사 정보 읽기
        반환: {회사명: {"고유번호": "...", "종목코드": "..."}}
        """
        df = pd.read_excel(filepath, sheet_name="Company", header=None)
        companies = {}
        
        for idx, row in df.iterrows():
            if idx == 0:  # 헤더 스킵
                continue
            
            company_name = str(row[0]).strip() if pd.notna(row[0]) else ""
            corp_id = str(row[1]).strip() if pd.notna(row[1]) else ""
            ticker = str(row[2]).strip() if pd.notna(row[2]) else ""
            
            if company_name:
                companies[company_name] = {
                    "고유번호": corp_id,
                    "종목코드": ticker
                }
        
        logger.info(f"회사 {len(companies)}개 로드 완료")
        return companies

    @staticmethod
    def read_keywords(filepath: str) -> List[Dict[str, str]]:
        """
        ESG 시트에서 키워드 읽기
        반환: [{"theme": "...", "key_issue": "...", "keyword": "..."}, ...]
        """
        df = pd.read_excel(filepath, sheet_name="ESG", header=None)
        keywords = []
        
        for idx, row in df.iterrows():
            if idx == 0:  # 헤더 스킵
                continue
            
            theme = str(row[0]).strip() if pd.notna(row[0]) else ""  # A열: Theme (주제)
            key_issue = str(row[2]).strip() if pd.notna(row[2]) else ""  # C열: Key Issue
            keyword_str = str(row[3]).strip() if pd.notna(row[3]) else ""  # D열: 뉴스 키워드 후보
            
            if keyword_str:
                # 쉼표로 분리
                for kw in keyword_str.split(","):
                    kw = kw.strip()
                    if kw:
                        keywords.append({
                            "theme": theme,
                            "key_issue": key_issue,
                            "keyword": kw
                        })
        
        logger.info(f"키워드 {len(keywords)}개 로드 완료")
        return keywords

    @staticmethod
    def save_results(results: List[Dict[str, Any]], output_path: str):
        """결과를 엑셀로 저장"""
        df = pd.DataFrame(results)
        
        # 컬럼 순서 정의
        column_order = [
            "esg", "Theme (주제)", "Key Issue (핵심 이슈)", "뉴스 키워드 후보",
            "부정 ESG 키워드", "부정점수", "뉴스 보도날짜", "기사제목",
            "언론사", "기사 URL", "회사명", "고유번호", "종목코드"
        ]
        
        # 누락된 컬럼 추가
        for col in column_order:
            if col not in df.columns:
                df[col] = ""
        
        df = df[column_order]
        
        # 저장
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="output", index=False)
            
            # 스타일 적용
            ws = writer.sheets["output"]
            ExcelHandler._apply_header_style(ws)
            ExcelHandler._autosize_columns(ws)
        
        logger.info(f"✅ 결과 저장 완료: {output_path}")
        print(f"SAVED: {Path(output_path).resolve()}")

    @staticmethod
    def _apply_header_style(ws):
        """헤더 스타일 적용"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    @staticmethod
    def _autosize_columns(ws, max_width: int = 50):
        """컬럼 너비 자동 조정"""
        for column in ws.columns:
            length = max((len(str(cell.value)) if cell.value else 0) for cell in column)
            col_letter = column[0].column_letter
            ws.column_dimensions[col_letter].width = min(length + 2, max_width)


# ========================================
# 3. 메인 크롤러 로직
# ========================================
class NewsCrawlerApp:
    """통합 크롤러 앱"""

    def __init__(self, input_file: str, output_dir: str = "./output"):
        self.input_file = input_file
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        self.crawler = NaverNewsCrawler()
        self.excel_handler = ExcelHandler()

    def run(self):
        """크롤링 실행"""
        logger.info("=" * 50)
        logger.info("네이버 뉴스 크롤링 시작")
        logger.info("=" * 50)
        
        # 1. 엑셀에서 데이터 읽기
        companies = self.excel_handler.read_companies(self.input_file)
        keywords_info = self.excel_handler.read_keywords(self.input_file)
        
        # 2. 크롤링 수행
        results = []
        total = len(companies) * len(keywords_info)
        current = 0
        
        for company_name, company_info in companies.items():
            for kw_info in keywords_info:
                current += 1
                theme = kw_info["theme"]
                keyword = kw_info["keyword"]
                key_issue = kw_info["key_issue"]
                
                query = f"{company_name} {keyword}"
                logger.info(f"[{current}/{total}] 검색: {query}")
                
                try:
                    # 뉴스 검색
                    raw_items = self.crawler.search_news_multiple_pages(query, max_results=100)
                    formatted_items = self.crawler.format_news_data(raw_items, keyword)
                    
                    # 결과 포맷팅
                    for item in formatted_items:
                        results.append({
                            "esg": "",  # 추후 분류
                            "Theme (주제)": theme,
                            "Key Issue (핵심 이슈)": key_issue,
                            "뉴스 키워드 후보": keyword,
                            "부정 ESG 키워드": "",
                            "부정점수": -1,
                            "뉴스 보도날짜": item["pub_date"],
                            "기사제목": item["title"],
                            "언론사": "",  # 네이버 API에서 제공 안 함
                            "기사 URL": item["link"],
                            "회사명": company_name,
                            "고유번호": company_info["고유번호"],
                            "종목코드": company_info["종목코드"],
                        })
                    
                    logger.info(f"  ✓ {len(formatted_items)}건 수집")
                    
                except Exception as e:
                    logger.error(f"  ✗ 오류: {e}")
                
                # API 호출 제한 대응
                time.sleep(0.2)
        
        # 3. 결과 저장
        if results:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = self.output_dir / f"news_output_{timestamp}.xlsx"
            self.excel_handler.save_results(results, str(output_file))
            logger.info(f"총 {len(results)}건의 뉴스 수집 완료")
        else:
            logger.warning("수집된 뉴스가 없습니다.")
        
        self.crawler.close()


# ========================================
# 4. 실행
# ========================================
if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("사용법: python crawl.py input.xlsx")
        sys.exit(1)
    
    input_file = sys.argv[1]
    
    if not Path(input_file).exists():
        print(f"파일을 찾을 수 없습니다: {input_file}")
        sys.exit(1)
    
    try:
        app = NewsCrawlerApp(input_file)
        app.run()
    except Exception as e:
        logger.error(f"실행 오류: {e}", exc_info=True)
        sys.exit(1)