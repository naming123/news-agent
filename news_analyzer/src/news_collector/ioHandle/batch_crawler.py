"""
엑셀 기반 입출력 처리 모듈 (B안: date_from/date_to 시그니처 확장)
"""
from __future__ import annotations
from datetime import datetime
from zoneinfo import ZoneInfo  # Python 3.9+
import logging
import re
from collections import defaultdict
from pathlib import Path
from datetime import datetime, date
from typing import List, Dict, Any, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from types import SimpleNamespace
from ..collector.crawler import NaverNewsCrawler
from ..config.models import NewsArticle, CrawlerConfig  # ← 누락 보강

logger = logging.getLogger(__name__)

# =========================
# Batch Crawler
# =========================
def now_kst(fmt="%Y-%m-%d %H:%M:%S"):
    return datetime.now(ZoneInfo("Asia/Seoul")).strftime(fmt)
class BatchNewsCrawler:
    def __init__(
        self,
        input: str,
        output_dir: str = "./output",
        config: dict | None = None,
        *,
        inplace: bool = False,
        output_sheet: str = "output",
        args=None
    ):
        self.input = input
        self.output_dir = Path(output_dir); self.output_dir.mkdir(parents=True, exist_ok=True)
        from types import SimpleNamespace
        self.args = args or SimpleNamespace(
            companies=None, keywords=None, start=None, end=None,
            output_sheet=None, inplace=False, max_results=None,
        )

        self.config = config or {}                # ← 이름도 명확히
        self.excel_config = self.config           # (아래 코드 호환되게 유지)

        # 크롤러 생성 시도
        self.crawler = NaverNewsCrawler(config=self.config)

        from collections import defaultdict
        self.results_by_keyword = defaultdict(list)
        self.results_by_company = defaultdict(dict)

        self.inplace = inplace
        self.output_sheet = output_sheet

    def run(self) -> Dict[str, List[NewsArticle | Dict[str, Any]]]:
        """
        - Company 시트 × ESG 키워드 조합으로 네이버 뉴스 검색
        - 기간 필터(date_from/date_to) 적용
        - 결과는 self.results_by_keyword 에 {keyword: [기사...]} 형태로 누적
        - 마지막에 별도 타임스탬프 파일로 저장하고, SAVED 로그를 출력
        """
   
        # --- 회사: CLI > Excel ---
        if self.args.companies is not None:
            companies = [s.strip() for s in str(self.args.companies).split(",") if s.strip()]
        else:
            companies = ExcelInputHandler.read_companies(self.input)

        # --- 키워드: CLI > Excel ---
        if self.args.keywords is not None:
            key_specs = [{"group": "", "keyword": s}
                         for s in str(self.args.keywords).split(",") if s.strip()]
        else:
            key_specs = ExcelInputHandler.read_keywords(self.input)

        # 만약 엑셀에서도 못 읽었다면 → CLI 인자 fallback
        if not key_specs and self.args.keywords:
            key_specs = [{"group": "", "keyword": s} for s in self.args.keywords.split(",") if s.strip()]


        # Config에서 기간 파라미터 도출(YYYY-MM-DD 형태로 정규화 시도)
        date_from_raw = self.args.start or self.excel_config.get("date_from")
        date_to_raw   = self.args.end   or self.excel_config.get("date_to")
        date_from = ExcelInputHandler._normalize_date_ymd(date_from_raw)
        date_to   = ExcelInputHandler._normalize_date_ymd(date_to_raw)

        for company in companies:
            for spec in key_specs:
                group = spec.get("group", "")
                kw    = spec.get("keyword", "")
                if not kw:
                    continue

                query = f"{company} {kw}".strip() if company else kw
                print(f"[CALL] query='{query}'  from={date_from} to={date_to}")
                try:
                    # 끝까지 혹은 기간 하한 도달 시까지 수집
                    raw_items = self.crawler.search_news_multiple_pages(
                        query=query,
                        max_results=None,             # None → 끝까지/기간조건
                        date_from=date_from,
                        date_to=date_to,
                    )
                    # dict 리스트를 프로젝트 표준 dict로 변환
                    formatted = self.crawler.format_news_data(raw_items, query)

                    # NewsArticle로 감싸도 되고(dict도 허용). 메타 주입
                    articles: List[NewsArticle | Dict[str, Any]] = []
                    for item in formatted:
                        a = {
                            **item,                      # title/description/link/pub_date...
                            "company": company,
                            "keyword": kw,
                            "group": group,
                            "search_query": query,
                            "date_from": date_from or "",
                            "date_to": date_to or "",
                        }
                        articles.append(a)

                    self.results_by_keyword[kw].extend(articles)
                    self.results_by_company.setdefault(company, {}).setdefault(kw, []).extend(articles)
                    print(f"[DBG] args.companies={self.args.companies!r}, args.keywords={self.args.keywords!r}")
                    print(f"[DBG] companies={companies}")
                    print(f"[DBG] keywords={[d.get('keyword') for d in key_specs]}  (n={len(key_specs)})")
                    print(f"[DBG] date_from={date_from} date_to={date_to}")
                except Exception as e:
                    logger.error(f"[{company}] {group}/{kw} 에러: {e}", exc_info=True)


        self._save_results(date_from=date_from, date_to=date_to)
        return self.results_by_keyword

    def _save_results(self, *, date_from: Optional[str], date_to: Optional[str]) -> None:
        """
        항상 별도 파일(output/news_output_타임스탬프.xlsx)에 저장.
        Streamlit 앱이 파싱할 수 있도록 'SAVED: <abs_path>' 한 줄 출력.
        """
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = self.output_dir / f"news_output_{ts}.xlsx"
        out_path.parent.mkdir(parents=True, exist_ok=True)

        ExcelOutputHandler.save_results_to_sheet(
            results_by_keyword=self.results_by_keyword,
            file_path=str(out_path),
            sheet_name="output",
            replace=True,
            date_from=date_from,
            date_to=date_to,
        )

        # 앱이 이 줄을 정규식으로 잡아 경로를 인식함
        print(f"SAVED: {out_path.resolve()}")
        print("final: "+now_kst())   

    def close(self):
        if self.crawler:
            self.crawler.close()


# =========================
# Input Handlers
# =========================
class ExcelInputHandler:
    """엑셀 입력 처리"""

    @staticmethod
    def read_companies(filepath: str, sheet_name: str = "Company", column: str = "A", start_row: int = 2) -> List[str]:
    
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
        col_idx = ord(column.upper()) - ord("A")
        return [str(v).strip() for v in df.iloc[start_row - 1:, col_idx].dropna() if str(v).strip()]

    @staticmethod
    def read_keywords(filepath: str, sheet_name: str = "ESG", base_col: str = "C", cand_col: str = "D", start_row: int = 2) -> List[Dict[str, Any]]:
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
        base_idx = ord(base_col.upper()) - ord("A")
        cand_idx = ord(cand_col.upper()) - ord("A")

        rows = df.iloc[start_row - 1:, [base_idx, cand_idx]].dropna(how="all")
        out = []
        for _, (base_val, cand_val) in rows.iterrows():
            base = str(base_val).strip() if pd.notna(base_val) else ""
            cands = str(cand_val).split(",") if pd.notna(cand_val) else []
            for kw in (k.strip() for k in cands):
                if kw:
                    out.append({"group": base, "keyword": kw, "company": ""})
        return out

    @staticmethod
    def _normalize_date_ymd(v: Optional[str]) -> Optional[str]:
        if not v:
            return None
        s = str(v).strip()
        for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d"):
            try:
                return datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
            except Exception:
                pass
        return None

    @staticmethod
    def read_config(filepath: str, sheet_name: str = "Config") -> Dict[str, Any]:
        defaults = {"max_pages": 3, "min_delay": 1.0, "max_delay": 2.0}
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name)
            config: Dict[str, Any] = {}
            for _, row in df.iterrows():
                if "parameter" in df.columns and "value" in df.columns:
                    param = str(row["parameter"]).strip()
                    value = row["value"]
                    if "type" in df.columns:
                        t = str(row.get("type", "str")).lower()
                        if t == "int":
                            value = int(value)
                        elif t == "float":
                            value = float(value)
                        elif t == "bool":
                            value = str(value).lower() in ["true", "1", "yes"]
                    config[param] = value

            # 연 단위 기간 처리 → date_from/date_to로 보정
            if "date_from" not in config and "date_to" not in config:
                if "year" in config:
                    y = int(config["year"])
                    config["date_from"] = f"{y}.01.01"
                    config["date_to"] = f"{y}.12.31"
                elif "start_year" in config and "end_year" in config:
                    sy, ey = int(config["start_year"]), int(config["end_year"])
                    config["date_from"] = f"{sy}.01.01"
                    config["date_to"] = f"{ey}.12.31"

            # 표준 포맷(YYYY-MM-DD)로 통일
            if config.get("date_from"):
                config["date_from"] = ExcelInputHandler._normalize_date_ymd(config["date_from"])
            if config.get("date_to"):
                config["date_to"] = ExcelInputHandler._normalize_date_ymd(config["date_to"])

            for k, v in defaults.items():
                config.setdefault(k, v)
            return config
        except Exception as e:
            logger.warning(f"Config sheet not found, using defaults: {e}")
            return defaults


# =========================
# Output Handlers (새 스키마)
# =========================
class ExcelOutputHandler:
    """엑셀 출력 처리 (새 스키마: ESG/Theme/Key Issue/...)"""

    @staticmethod
    def _safe_get(obj: Any, key: str, default: Any = "") -> Any:
        if isinstance(obj, dict):
            return obj.get(key, default)
        return getattr(obj, key, default)

    @staticmethod
    def _derive_esg(group: str, fallback_esg: str = "") -> str:
        """group이나 esg필드로 E/S/G/F 추론 (정확 분류가 있으면 그대로 사용)"""
        if fallback_esg:
            return str(fallback_esg).strip()
        g = (group or "").strip().upper()
        if g.startswith("E"): return "E"
        if g.startswith("S"): return "S"
        if g.startswith("G"): return "G"
        if g.startswith("F") or "KOSELF" in g: return "F"
        return ""

    @staticmethod
    def _yyyymmdd(any_date) -> str:
        """다양한 입력 → YYYYMMDD 문자열"""
        if any_date in (None, ""):
            return ""
        if isinstance(any_date, datetime):
            return any_date.strftime("%Y%m%d")
        if isinstance(any_date, date):
            return any_date.strftime("%Y%m%d")
        s = str(any_date)
        for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d"):
            try:
                return datetime.strptime(s[:10], fmt).strftime("%Y%m%d")
            except Exception:
                pass
        try:
            from email.utils import parsedate_to_datetime
            return parsedate_to_datetime(s).strftime("%Y%m%d")
        except Exception:
            return s

    @staticmethod
    def save_results_to_sheet(
        results_by_keyword: Dict[str, List[Any]],
        file_path: str,
        sheet_name: str = "output",
        replace: bool = True,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
    ) -> None:
        """
        결과를 'output' 시트에 새 스키마로 저장한다.
        새 스키마 컬럼:
        [esg, Theme (주제), Key Issue (핵심 이슈), 뉴스 키워드 후보, 부정 ESG 키워드, 부정점수,
         뉴스 보도날짜(YYYYMMDD), 기사제목, 언론사, 기사 URL, 회사명, 고유번호, 종목코드]
        - 부정점수는 요청대로 1로 고정
        """
        rows: List[dict] = []
        sg = ExcelOutputHandler._safe_get
        for kw, items in (results_by_keyword or {}).items():
            for a in (items or []):
                esg        = sg(a, "esg", "")
                group      = sg(a, "group", "")
                theme      = sg(a, "theme", group)
                key_issue  = sg(a, "key_issue", sg(a, "keyword", kw))
                news_kw    = sg(a, "keyword", kw)
                neg_terms  = sg(a, "neg_terms", "")
                title      = sg(a, "title", "")
                press      = sg(a, "press", sg(a, "source", ""))
                url        = sg(a, "link", sg(a, "url", ""))
                company    = sg(a, "company", "")
                corp_id    = sg(a, "corp_id", "")
                ticker     = sg(a, "ticker", "")
                raw_date   = sg(a, "date", "") or sg(a, "pub_date", "") or sg(a, "pubDate", "")
                ymd        = ExcelOutputHandler._yyyymmdd(raw_date)

                if isinstance(neg_terms, (list, tuple, set)):
                    neg_terms = ", ".join(map(str, neg_terms))

                esg_final = ExcelOutputHandler._derive_esg(group, fallback_esg=str(esg).strip())

                rows.append({
                    "esg": esg_final,
                    "Theme (주제)": theme,
                    "Key Issue (핵심 이슈)": key_issue,
                    "뉴스 키워드 후보": news_kw,
                    "부정 ESG 키워드": neg_terms,
                    "부정점수": 1,  # ← 요청대로 1 고정
                    "뉴스 보도날짜(YYYYMMDD)": ymd,
                    "기사제목": title,
                    "언론사": press,
                    "기사 URL": url,
                    "회사명": company,
                    "고유번호": corp_id,
                    "종목코드": ticker,
                })

        df = pd.DataFrame(rows)
        # 열 순서 보장 및 누락 보정
        desired = [
            "esg", "Theme (주제)", "Key Issue (핵심 이슈)", "뉴스 키워드 후보",
            "부정 ESG 키워드", "부정점수", "뉴스 보도날짜(YYYYMMDD)",
            "기사제목", "언론사", "기사 URL", "회사명", "고유번호", "종목코드",
        ]
        for c in desired:
            if c not in df.columns:
                df[c] = "" if c != "부정점수" else 1
        df = df[desired]

        p = Path(file_path)
        p.parent.mkdir(parents=True, exist_ok=True)

        if p.exists():
            book = load_workbook(p)
            if sheet_name in book.sheetnames and replace:
                book.remove(book[sheet_name])
            with pd.ExcelWriter(p, engine="openpyxl", mode="a") as writer:
                writer.book = book
                df.to_excel(writer, index=False, sheet_name=sheet_name)

                meta_df = pd.DataFrame(
                    [
                        {"key": "updated_at", "value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
                        {"key": "date_from",  "value": date_from or ""},
                        {"key": "date_to",    "value": date_to or ""},
                        {"key": "total_rows", "value": len(df)},
                        {"key": "sheet_name", "value": sheet_name},
                    ]
                )
                if "meta" in writer.book.sheetnames and replace:
                    writer.book.remove(writer.book["meta"])
                meta_df.to_excel(writer, index=False, sheet_name="meta")
                writer.save()
        else:
            with pd.ExcelWriter(p, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                meta_df = pd.DataFrame(
                    [
                        {"key": "created_at", "value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
                        {"key": "date_from",  "value": date_from or ""},
                        {"key": "date_to",    "value": date_to or ""},
                        {"key": "total_rows", "value": len(df)},
                        {"key": "sheet_name", "value": sheet_name},
                    ]
                )
                meta_df.to_excel(writer, index=False, sheet_name="meta")

    # ------- (아래 보조들: 스타일/유틸) -------
    @staticmethod
    def _apply_summary_style(ws) -> None:
        try:
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            ExcelOutputHandler._autosize_columns(ws)
        except Exception as e:
            logger.warning(f"Could not apply styles: {str(e)}")

    @staticmethod
    def _autosize_columns(ws, max_width: int = 50):
        for column in ws.columns:
            length = max((len(str(cell.value)) if cell.value else 0) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(length + 2, max_width)


# =========================
# 편의 유틸 (선택)
# =========================
def save_keyword_results(
    articles_by_keyword: Dict[str, List[NewsArticle]],
    output_dir: str,
    filename_prefix: str = "keyword_results",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
) -> str:
    """
    크롤링 결과를 새 스키마로 저장하고 파일 경로 반환
    """
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = output_path / f"{filename_prefix}_{timestamp}.xlsx"

    ExcelOutputHandler.save_results_to_sheet(
        results_by_keyword=articles_by_keyword,
        file_path=str(filepath),
        sheet_name="output",
        replace=True,
        date_from=date_from,
        date_to=date_to,
    )
    return str(filepath)
# 파일 끝에 추가
if __name__ == "__main__":
    print("batch: "+now_kst()) 
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--input")
    parser.add_argument("--companies", type=str)
    parser.add_argument("--keywords", type=str)
    parser.add_argument("--start", type=str)
    parser.add_argument("--end", type=str)
    parser.add_argument("--output-sheet", default="output")
    parser.add_argument("--inplace", action="store_true")

    
    args = parser.parse_args()
    
    def _norm_ymd(s: str | None) -> str | None:
        if not s: return None
        from datetime import datetime
        for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d"):
            try:
                return datetime.strptime(str(s)[:10], fmt).strftime("%Y-%m-%d")
            except Exception:
                pass
        return None

    cli_config = {
        # 필요시 추가할 수 있는 값들 (딜레이/타임아웃 등도 CLI에 넣으면 같이 넘기자)
        "date_from": _norm_ymd(getattr(args, "start", None)),
        "date_to":   _norm_ymd(getattr(args, "end",   None)),
        # "min_delay": float(getattr(args, "min_delay", 1.0)),
        # "max_delay": float(getattr(args, "max_delay", 2.0)),
        # "timeout":   float(getattr(args, "timeout", 10.0)),
        # "max_page":  int(getattr(args, "max_page", 3)),
    }

    app = BatchNewsCrawler(
        input=args.input,
        config=cli_config,     # ← 엑셀 대신 CLI 구성 딕셔너리 전달
        args=args,
    )
    app.run()
    app.close()