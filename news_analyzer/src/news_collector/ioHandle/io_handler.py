"""
엑셀 기반 입출력 처리 모듈
"""
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional
import re
from openpyxl import load_workbook
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ..config.models import NewsArticle

logger = logging.getLogger(__name__)

def _article_to_row(a: Any, fallback_keyword: str = "") -> dict:
    """
    NewsArticle(객체)든 dict든 안전하게 행으로 변환.
    프로젝트마다 필드명이 다를 수 있어 최대한 넓게 커버.
    """
    # dict/객체 양쪽 케이스 방어
    def g(obj, name, default=""):
        if isinstance(obj, dict):
            return obj.get(name, default)
        return getattr(obj, name, default)

    return {
        "회사":           g(a, "company"),
        "키워드":         g(a, "keyword", fallback_keyword),
        "그룹":           g(a, "group"),
        "제목":           g(a, "title"),
        "링크":           g(a, "link") or g(a, "url"),
        "원문링크":       g(a, "original_link") or g(a, "originallink"),
        "언론사":         g(a, "press") or g(a, "source"),
        "날짜":           g(a, "date") or g(a, "pub_date") or g(a, "pubDate"),
        "요약":           g(a, "summary") or g(a, "description"),
        "검색쿼리":       g(a, "search_query"),
        "수집시각":       g(a, "crawl_time") or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "date_from":      g(a, "date_from", ""),
        "date_to":        g(a, "date_to", ""),
    }




class ExcelInputHandler:
    """엑셀 입력 처리"""

    @staticmethod
    def read_companies(filepath: str, sheet_name: str = "Company", column: str = "A", start_row: int = 2) -> List[str]:
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
        col_idx = ord(column.upper()) - ord("A")
        return [str(v).strip() for v in df.iloc[start_row - 1:, col_idx].dropna() if str(v).strip()]

    @staticmethod
    def read_keywords(filepath: str, sheet_name: str = "ESG", base_col: str = "C", cand_col: str = "D", start_row: int = 2) -> List[Dict[str, Any]]:
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
        base_idx = ord(base_col.upper()) - ord("A")
        cand_idx = ord(cand_col.upper()) - ord("A")

        rows = df.iloc[start_row - 1:, [base_idx, cand_idx]].dropna(how="all")
        out = []
        for _, (base_val, cand_val) in rows.iterrows():
            base = str(base_val).strip() if pd.notna(base_val) else ""
            cands = str(cand_val).split(",") if pd.notna(cand_val) else []
            for kw in (k.strip() for k in cands):
                if kw:
                    out.append({"group": base, "keyword": kw, "company": ""})
        return out

    @staticmethod
    def read_config(filepath: str, sheet_name: str = "Config") -> Dict[str, Any]:
        defaults = {"max_pages": 3, "min_delay": 1.0, "max_delay": 2.0}
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name)
            config = {}
            for _, row in df.iterrows():
                if "parameter" in df.columns and "value" in df.columns:
                    param = str(row["parameter"]).strip()
                    value = row["value"]
                    if "type" in df.columns:
                        t = str(row.get("type", "str")).lower()
                        if t == "int": value = int(value)
                        elif t == "float": value = float(value)
                        elif t == "bool": value = str(value).lower() in ["true", "1", "yes"]
                    config[param] = value

            # 연 단위 기간 처리
            if "date_from" not in config and "date_to" not in config:
                if "year" in config:
                    y = int(config["year"])
                    config["date_from"] = f"{y}.01.01"
                    config["date_to"] = f"{y}.12.31"
                elif "start_year" in config and "end_year" in config:
                    sy, ey = int(config["start_year"]), int(config["end_year"])
                    config["date_from"] = f"{sy}.01.01"
                    config["date_to"] = f"{ey}.12.31"
            for k, v in defaults.items():
                config.setdefault(k, v)
            return config
        except Exception as e:
            logger.warning(f"Config sheet not found, using defaults: {e}")
            return defaults


class ExcelOutputHandler:
    """엑셀 출력 처리"""

    @staticmethod
    def save_results(articles_by_keyword: Dict[str, List[Any]], filepath: str):
        Path(filepath).parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            all_articles = []
            for kw, articles in articles_by_keyword.items():
                for a in articles:
                    row = a.to_dict() if hasattr(a, "to_dict") else {}
                    row["검색_회사"]  = getattr(a, "company", "")
                    row["기준키워드"] = getattr(a, "group", "")
                    row["후보키워드"] = getattr(a, "keyword", "")
                    row["검색_쿼리"]  = getattr(a, "search_query", "")
                    row["기간_from"] = getattr(a, "date_from", "")
                    row["기간_to"]   = getattr(a, "date_to", "")
                    all_articles.append(row)

            df = pd.DataFrame(all_articles)
            df.to_excel(writer, sheet_name="전체_데이터", index=False)

            # 스타일 적용
            try:
                ws = writer.sheets["전체_데이터"]
                ExcelOutputHandler._autosize_columns(ws)
            except Exception as e:
                logger.error(f"[save_results] 컬럼 자동 너비 조정 실패: {e}", exc_info=True)

    @staticmethod
    def save_company_results(
        results_by_company: Dict[str, Dict[str, List[NewsArticle]]],
        filepath: str
    ) -> None:
        """회사별 결과 저장 (호환성 유지)"""
        # 회사별 → 키워드별로 flatten
        all_articles = {}
        for company, keyword_articles in results_by_company.items():
            for keyword, articles in keyword_articles.items():
                key = f"{company}_{keyword}" if company else keyword
                all_articles[key] = articles
        
        ExcelOutputHandler.save_results(all_articles, filepath)

    @staticmethod
    def _clean_sheet_name(name: str, max_length: int = 31) -> str:
        """엑셀 시트명 정리"""
        # 특수문자 제거
        invalid = r'[\/\\\?\*\[\]\:]'
        safe = re.sub(invalid, "_", str(name)).strip()
        
        # 길이 제한
        if len(safe) > max_length:
            safe = safe[:max_length - 3] + "..."
        
        return safe or "Sheet"

    @staticmethod
    def _dedupe_sheet_name(base: str, used: set, max_length: int = 31) -> str:
        """중복 시트명 처리"""
        if base not in used:
            return base
        
        i = 2
        while True:
            suffix = f"_{i}"
            # 충분한 공간 확보
            if len(base) + len(suffix) <= max_length:
                candidate = base + suffix
            else:
                candidate = base[:max_length - len(suffix)] + suffix
            
            if candidate not in used:
                return candidate
            i += 1

    @staticmethod
    def _apply_summary_style(ws) -> None:
        """요약 시트 스타일 적용"""
        try:
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 헤더 행 스타일
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            ExcelOutputHandler._autosize_columns(ws)
            
        except Exception as e:
            logger.warning(f"Could not apply styles: {str(e)}")
    
    @staticmethod
    def save_results_to_sheet(
        results_by_keyword: dict,
        file_path: str,
        sheet_name: str = "output",
        replace: bool = True,
    ):
        """
        결과를 'output' 시트(필요 시 output_2, output_3...)로 분할 저장한다.
        새 스키마 컬럼:
        [esg, Theme (주제), Key Issue (핵심 이슈), 뉴스 키워드 후보, 부정 ESG 키워드, 부정점수,
         뉴스 보도날짜(YYYYMMDD), 기사제목, 언론사, 기사 URL, 회사명, 고유번호, 종목코드]
        """
        import math
        from datetime import datetime, date
        from openpyxl import load_workbook
        import pandas as pd
        import numpy as np
        from pathlib import Path

        EXCEL_MAX_ROWS = 1_048_576  # 엑셀 시트 최대 행수

        def _safe_get(obj, key, default=""):
            if isinstance(obj, dict):
                return obj.get(key, default)
            return getattr(obj, key, default)

        def _derive_esg(group: str, fallback_esg: str = "") -> str:
            """group이나 esg필드로 E/S/G/F 추론"""
            if fallback_esg:
                return str(fallback_esg).strip()
            g = (group or "").strip().upper()
            if g.startswith("E"): return "E"
            if g.startswith("S"): return "S"
            if g.startswith("G"): return "G"
            if "KOSELF" in g or g.startswith("F"): return "F"
            return ""

        def _yyyymmdd(any_date) -> str:
            """여러 형식 → YYYYMMDD 문자열"""
            if any_date in (None, ""):
                return ""
            if isinstance(any_date, datetime):
                return any_date.strftime("%Y%m%d")
            if isinstance(any_date, date):
                return any_date.strftime("%Y%m%d")
            s = str(any_date)
            for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d"):
                try:
                    return datetime.strptime(s[:10], fmt).strftime("%Y%m%d")
                except Exception:
                    pass
            try:
                from email.utils import parsedate_to_datetime
                dt = parsedate_to_datetime(s)
                return dt.strftime("%Y%m%d")
            except Exception:
                return s

        # 1) rows 생성 (스키마 정규화)
        rows = []
        for kw, articles in (results_by_keyword or {}).items():
            for a in (articles or []):
                esg           = _safe_get(a, "esg", "")
                group         = _safe_get(a, "group", "")
                theme         = _safe_get(a, "theme", group)
                key_issue     = _safe_get(a, "key_issue", _safe_get(a, "keyword", kw))
                news_kw_cand  = _safe_get(a, "keyword", kw)
                neg_terms     = _safe_get(a, "neg_terms", "")
                neg_score     = _safe_get(a, "neg_score", 0)
                title         = _safe_get(a, "title", "")
                press         = _safe_get(a, "press", _safe_get(a, "source", ""))
                url           = _safe_get(a, "link", _safe_get(a, "url", ""))
                company       = _safe_get(a, "company", "")
                corp_id       = _safe_get(a, "corp_id", "")
                ticker        = _safe_get(a, "ticker", "")
                raw_date      = (_safe_get(a, "date", "")
                                 or _safe_get(a, "pub_date", "")
                                 or _safe_get(a, "pubDate", ""))

                # 타입 보정
                if isinstance(neg_terms, (list, tuple, set)):
                    neg_terms = ", ".join(map(str, neg_terms))
                try:
                    # 점수는 float로 강제
                    neg_score = float(neg_score)
                except Exception:
                    neg_score = 0.0

                ymd = _yyyymmdd(raw_date)
                esg_final = _derive_esg(group, fallback_esg=str(esg).strip())

                rows.append({
                    "esg": esg_final,
                    "Theme (주제)": theme,
                    "Key Issue (핵심 이슈)": key_issue,
                    "뉴스 키워드 후보": news_kw_cand,
                    "부정 ESG 키워드": neg_terms,
                    "부정점수": neg_score,  # ✅ 실제 점수 기록
                    "뉴스 보도날짜(YYYYMMDD)": ymd,
                    "기사제목": title,
                    "언론사": press,
                    "기사 URL": url,
                    "회사명": company,
                    "고유번호": corp_id,
                    "종목코드": ticker,
                })

        df = pd.DataFrame(rows)

        # 2) 열 순서/결측 보정
        desired_order = [
            "esg", "Theme (주제)", "Key Issue (핵심 이슈)", "뉴스 키워드 후보",
            "부정 ESG 키워드", "부정점수", "뉴스 보도날짜(YYYYMMDD)",
            "기사제목", "언론사", "기사 URL", "회사명", "고유번호", "종목코드",
        ]
        for col in desired_order:
            if col not in df.columns:
                # 스키마 유지: 점수는 0.0, 나머지는 빈 문자열
                df[col] = 0.0 if col == "부정점수" else ""
        df = df[desired_order]

        # 3) 저장 경로/디렉토리
        target = Path(file_path)
        target.parent.mkdir(parents=True, exist_ok=True)

        # 4) Excel 저장: 행 제한 초과 시 자동 분할(output, output_2, ...)
        def _remove_existing_output_sheets(book, base_name: str):
            """replace=True일 때 기존 output, output_2.. 시트를 삭제."""
            to_remove = []
            for s in list(book.sheetnames):
                if s == base_name or (s.startswith(base_name + "_") and s[len(base_name)+1:].isdigit()):
                    to_remove.append(s)
            for s in to_remove:
                ws = book[s]
                book.remove(ws)

        def _write_chunks(writer, df_chunked, base_name: str):
            chunks = math.ceil(len(df_chunked) / EXCEL_MAX_ROWS) if len(df_chunked) else 1
            if len(df_chunked) == 0:
                # 최소 1시트 보장 (빈 데이터라도)
                pd.DataFrame({"info": ["no rows"]}).to_excel(writer, index=False, sheet_name=base_name)
                return
            for i in range(chunks):
                start = i * EXCEL_MAX_ROWS
                end = min((i + 1) * EXCEL_MAX_ROWS, len(df_chunked))
                chunk_df = df_chunked.iloc[start:end]
                sname = base_name if i == 0 else f"{base_name}_{i+1}"
                chunk_df.to_excel(writer, index=False, sheet_name=sname)
                # 스타일 훅이 있으면 적용
                try:
                    ws = writer.book[sname]
                    ExcelOutputHandler._apply_summary_style(ws)  # 기존 헬퍼 유지
                except Exception:
                    pass

        # 5) 파일 존재 여부에 따라 저장
        try:
            # 기존 파일 존재 시
            book = load_workbook(str(target))
            if replace:
                _remove_existing_output_sheets(book, sheet_name)
            with pd.ExcelWriter(str(target), engine="openpyxl", mode="a") as writer:
                writer.book = book
                _write_chunks(writer, df, sheet_name)
                writer.save()
        except FileNotFoundError:
            # 새 파일 생성
            with pd.ExcelWriter(str(target), engine="openpyxl") as writer:
                _write_chunks(writer, df, sheet_name)
                writer.save()

    @staticmethod
    def _autosize_columns(ws, max_width: int = 50):
        for column in ws.columns:
            length = max((len(str(cell.value)) if cell.value else 0) for cell in column)
            col_letter = column[0].column_letter
            ws.column_dimensions[col_letter].width = min(length + 2, max_width)


    def save_keyword_results(
        articles_by_keyword: Dict[str, List[NewsArticle]], 
        output_dir: str, 
        filename_prefix: str = "keyword_results"
    ) -> str:
        """
        크롤링 결과를 엑셀로 저장하고 경로 반환
        """
        # 출력 디렉토리 생성
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        # 파일명 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = output_path / f"{filename_prefix}_{timestamp}.xlsx"
        
        # 저장
        ExcelOutputHandler.save_results(
            articles_by_keyword=articles_by_keyword,
            filepath=str(filepath),
            summary_sheet=True,
            separate_sheets=True
        )
        
        return str(filepath)