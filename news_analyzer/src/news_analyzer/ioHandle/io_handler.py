"""
엑셀 기반 입출력 처리 모듈
"""
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional
import re

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from news_analyzer.config.models import NewsArticle

logger = logging.getLogger(__name__)


class ExcelInputHandler:
    """엑셀 입력 처리"""

    @staticmethod
    def read_keywords(
        filepath: str,
        sheet_name: str = 'ESG',
        column: str = 'D',
        start_row: int = 2
    ) -> List[Dict[str, Any]]:
        """
        ESG 시트에서 키워드 읽기 (D열, 2행부터, 콤마로 분리)
        """
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
            
            # D열만 추출 (Excel D → pandas index 3)
            col_idx = ord(column.upper()) - ord('A')
            keywords_raw = df.iloc[start_row - 1:, col_idx].dropna()
            
            keywords = []
            for cell in keywords_raw:
                for kw in str(cell).split(','):
                    kw = kw.strip()
                    if kw:
                        keywords.append({
                            'keyword': kw,
                            'company': '',  # 회사명 없음
                            'metadata': {}
                        })
            
            logger.info(f"Loaded {len(keywords)} keywords from {filepath} [{sheet_name}!{column}{start_row}:]")
            return keywords
            
        except Exception as e:
            logger.error(f"Error reading ESG keywords: {str(e)}")
            raise

    @staticmethod
    def read_config(filepath: str, sheet_name: str = 'Config') -> Dict[str, Any]:
        """Config 시트에서 설정 읽기 (없으면 기본값 반환)"""
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name)
            config = {}
            for _, row in df.iterrows():
                if 'parameter' in df.columns and 'value' in df.columns:
                    param = str(row['parameter']).strip()
                    value = row['value']
                    
                    if 'type' in df.columns:
                        type_str = str(row.get('type', 'str')).lower()
                        if type_str == 'int':
                            value = int(value)
                        elif type_str == 'float':
                            value = float(value)
                        elif type_str == 'bool':
                            value = str(value).lower() in ['true', '1', 'yes']
                    
                    config[param] = value
            return config
        except Exception as e:
            logger.warning(f"Config sheet not found: {str(e)}. Using default values.")
            return {'max_pages': 3, 'min_delay': 1.0, 'max_delay': 2.0}


class ExcelOutputHandler:
    """엑셀 출력 처리"""

    @staticmethod
    def save_results(
        articles_by_keyword: Dict[str, List[NewsArticle]],
        filepath: str,
        summary_sheet: bool = True,
        separate_sheets: bool = True
    ) -> None:
        """크롤링 결과를 엑셀로 저장"""
        try:
            with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
                
                # 1. 요약 시트
                if summary_sheet:
                    summary_data = []
                    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    
                    for keyword, articles in articles_by_keyword.items():
                        latest_title = ""
                        if articles:
                            latest_title = articles[0].title if articles[0].title else ""
                        
                        press_count = len({a.press for a in articles if a.press})
                        
                        summary_data.append({
                            "키워드": keyword,
                            "수집 기사 수": len(articles),
                            "수집 시각": now_str,
                            "언론사 수": press_count,
                            "최신 기사": latest_title,
                        })
                    
                    if summary_data:
                        df = pd.DataFrame(summary_data)
                        df.to_excel(writer, sheet_name="요약", index=False)
                        
                        ws = writer.sheets["요약"]
                        ExcelOutputHandler._apply_summary_style(ws)

                # 2. 전체 데이터 시트
                all_rows = []
                for keyword, articles in articles_by_keyword.items():
                    for article in articles:
                        row = article.to_dict()
                        row["검색_키워드"] = keyword
                        all_rows.append(row)

                if all_rows:
                    all_df = pd.DataFrame(all_rows)
                    # 컬럼 순서 조정
                    preferred = ["검색_키워드", "title", "press", "date", "link", "crawled_at"]
                    cols = [c for c in preferred if c in all_df.columns]
                    remaining = [c for c in all_df.columns if c not in preferred]
                    all_df = all_df[cols + remaining]
                    
                    all_df.to_excel(writer, sheet_name="전체_데이터", index=False)
                    
                    ws_all = writer.sheets["전체_데이터"]
                    ExcelOutputHandler._autosize_columns(ws_all)

                # 3. 키워드별 개별 시트
                if separate_sheets:
                    sheet_names_used = set(["요약", "전체_데이터"])
                    
                    for keyword, articles in articles_by_keyword.items():
                        if not articles:
                            continue
                        
                        df = pd.DataFrame([a.to_dict() for a in articles])
                        preferred = ["title", "press", "date", "link", "crawled_at"]
                        cols = [c for c in preferred if c in df.columns]
                        remaining = [c for c in df.columns if c not in preferred]
                        df = df[cols + remaining]
                        
                        # 시트 이름 정리
                        sheet_name = ExcelOutputHandler._clean_sheet_name(keyword)
                        sheet_name = ExcelOutputHandler._dedupe_sheet_name(sheet_name, sheet_names_used)
                        sheet_names_used.add(sheet_name)
                        
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        ws = writer.sheets[sheet_name]
                        ExcelOutputHandler._autosize_columns(ws)

            logger.info(f"Results saved to {filepath}")
            
        except Exception as e:
            logger.error(f"Error saving results to Excel: {str(e)}")
            raise  # Exception 그대로 재발생

    @staticmethod
    def save_company_results(
        results_by_company: Dict[str, Dict[str, List[NewsArticle]]],
        filepath: str
    ) -> None:
        """회사별 결과 저장 (호환성 유지)"""
        # 회사별 → 키워드별로 flatten
        all_articles = {}
        for company, keyword_articles in results_by_company.items():
            for keyword, articles in keyword_articles.items():
                key = f"{company}_{keyword}" if company else keyword
                all_articles[key] = articles
        
        ExcelOutputHandler.save_results(all_articles, filepath)

    @staticmethod
    def _clean_sheet_name(name: str, max_length: int = 31) -> str:
        """엑셀 시트명 정리"""
        # 특수문자 제거
        invalid = r'[\/\\\?\*\[\]\:]'
        safe = re.sub(invalid, "_", str(name)).strip()
        
        # 길이 제한
        if len(safe) > max_length:
            safe = safe[:max_length - 3] + "..."
        
        return safe or "Sheet"

    @staticmethod
    def _dedupe_sheet_name(base: str, used: set, max_length: int = 31) -> str:
        """중복 시트명 처리"""
        if base not in used:
            return base
        
        i = 2
        while True:
            suffix = f"_{i}"
            # 충분한 공간 확보
            if len(base) + len(suffix) <= max_length:
                candidate = base + suffix
            else:
                candidate = base[:max_length - len(suffix)] + suffix
            
            if candidate not in used:
                return candidate
            i += 1

    @staticmethod
    def _apply_summary_style(ws) -> None:
        """요약 시트 스타일 적용"""
        try:
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 헤더 행 스타일
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            ExcelOutputHandler._autosize_columns(ws)
            
        except Exception as e:
            logger.warning(f"Could not apply styles: {str(e)}")

    @staticmethod
    def _autosize_columns(ws, max_width: int = 50) -> None:
        """컬럼 너비 자동 조정"""
        try:
            for column_cells in ws.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                adjusted_width = min(length + 2, max_width)
                
                # 컬럼 레터 가져오기
                column_letter = column_cells[0].column_letter
                ws.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.warning(f"Could not autosize columns: {str(e)}")


def save_keyword_results(
    articles_by_keyword: Dict[str, List[NewsArticle]], 
    output_dir: str, 
    filename_prefix: str = "keyword_results"
) -> str:
    """
    크롤링 결과를 엑셀로 저장하고 경로 반환
    """
    # 출력 디렉토리 생성
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    # 파일명 생성
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = output_path / f"{filename_prefix}_{timestamp}.xlsx"
    
    # 저장
    ExcelOutputHandler.save_results(
        articles_by_keyword=articles_by_keyword,
        filepath=str(filepath),
        summary_sheet=True,
        separate_sheets=True
    )
    
    return str(filepath)