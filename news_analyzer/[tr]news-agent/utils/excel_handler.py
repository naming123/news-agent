# utils/excel_handler.py
"""엑셀 기반 입출력 처리 모듈"""
from __future__ import annotations

import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def _article_to_row(a: Any, fallback_keyword: str = "") -> dict:
    """
    NewsArticle(객체)든 dict든 안전하게 행으로 변환
    """
    def g(obj, name, default=""):
        if isinstance(obj, dict):
            return obj.get(name, default)
        return getattr(obj, name, default)

    return {
        "회사":       g(a, "company", ""),
        "키워드":     g(a, "keyword", fallback_keyword),
        "그룹":       g(a, "group", ""),
        "제목":       g(a, "title", ""),
        "링크":       g(a, "link", "") or g(a, "url", ""),
        "원문링크":   g(a, "original_link", "") or g(a, "originallink", ""),
        "언론사":     g(a, "press", "") or g(a, "source", ""),
        "날짜":       g(a, "date", "") or g(a, "pub_date", "") or g(a, "pubDate", ""),
        "요약":       g(a, "summary", "") or g(a, "description", ""),
        "검색쿼리":   g(a, "search_query", ""),
        "수집시각":   g(a, "crawl_time", "") or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "date_from":  g(a, "date_from", ""),
        "date_to":    g(a, "date_to", ""),
    }


class ExcelInputHandler:
    """엑셀 입력 처리 클래스"""

    @staticmethod
    def read_companies(
        filepath: str, 
        sheet_name: str = "Company", 
        column: str = "A", 
        start_row: int = 2
    ) -> List[str]:
        """회사 목록 읽기"""
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
            col_idx = ord(column.upper()) - ord("A")
            companies = []
            
            for v in df.iloc[start_row - 1:, col_idx].dropna():
                company = str(v).strip()
                if company:
                    companies.append(company)
                    
            logger.info(f"Loaded {len(companies)} companies from {filepath}")
            return companies
            
        except Exception as e:
            logger.error(f"Failed to read companies: {e}")
            return []

    @staticmethod
    def read_keywords(
        filepath: str, 
        sheet_name: str = "ESG", 
        base_col: str = "C", 
        cand_col: str = "D", 
        start_row: int = 2
    ) -> List[Dict[str, Any]]:
        """키워드 목록 읽기"""
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
            base_idx = ord(base_col.upper()) - ord("A")
            cand_idx = ord(cand_col.upper()) - ord("A")

            rows = df.iloc[start_row - 1:, [base_idx, cand_idx]].dropna(how="all")
            keywords = []
            
            for _, (base_val, cand_val) in rows.iterrows():
                base = str(base_val).strip() if pd.notna(base_val) else ""
                cands = str(cand_val).split(",") if pd.notna(cand_val) else []
                
                for kw in cands:
                    kw = kw.strip()
                    if kw:
                        keywords.append({
                            "group": base, 
                            "keyword": kw, 
                            "company": ""
                        })
                        
            logger.info(f"Loaded {len(keywords)} keywords from {filepath}")
            return keywords
            
        except Exception as e:
            logger.error(f"Failed to read keywords: {e}")
            return []

    @staticmethod
    def read_config(filepath: str, sheet_name: str = "Config") -> Dict[str, Any]:
        """설정 시트 읽기"""
        defaults = {
            "max_pages": 3,
            "min_delay": 1.0,
            "max_delay": 2.0,
            "max_articles": 100
        }
        
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name)
            config = {}
            
            for _, row in df.iterrows():
                if "parameter" in df.columns and "value" in df.columns:
                    param = str(row["parameter"]).strip().lower()
                    value = row["value"]
                    
                    # 타입 변환
                    if "type" in df.columns:
                        t = str(row.get("type", "str")).lower()
                        if t == "int":
                            value = int(value)
                        elif t == "float":
                            value = float(value)
                        elif t == "bool":
                            value = str(value).lower() in ["true", "1", "yes"]
                    
                    config[param] = value

            # 날짜 처리
            if "date_from" not in config and "date_to" not in config:
                if "year" in config:
                    y = int(config["year"])
                    config["date_from"] = f"{y}.01.01"
                    config["date_to"] = f"{y}.12.31"
                elif "start_year" in config and "end_year" in config:
                    sy, ey = int(config["start_year"]), int(config["end_year"])
                    config["date_from"] = f"{sy}.01.01"
                    config["date_to"] = f"{ey}.12.31"

            # 기본값 적용
            for k, v in defaults.items():
                config.setdefault(k, v)
                
            logger.info(f"Loaded config: {config}")
            return config
            
        except Exception as e:
            logger.warning(f"Config sheet not found, using defaults: {e}")
            return defaults


class ExcelOutputHandler:
    """엑셀 출력 처리 클래스"""

    @staticmethod
    def save_results(
        results_by_keyword: Dict[str, List[Any]],
        file_path: str,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
    ) -> None:
        """
        키워드별 결과를 엑셀로 저장 (각 키워드는 별도 시트)
        """
        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        try:
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                # 메타 정보 시트
                meta_data = [
                    {"key": "exported_at", "value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
                    {"key": "date_from", "value": date_from or ""},
                    {"key": "date_to", "value": date_to or ""},
                    {"key": "total_keywords", "value": len(results_by_keyword)},
                    {"key": "keywords", "value": ", ".join(results_by_keyword.keys())},
                    {"key": "file_path", "value": str(path.resolve())},
                ]
                meta_df = pd.DataFrame(meta_data)
                meta_df.to_excel(writer, index=False, sheet_name="meta")

                # 결과가 없는 경우
                if not results_by_keyword:
                    empty_df = pd.DataFrame(columns=list(_article_to_row({}, "").keys()))
                    empty_df.to_excel(writer, index=False, sheet_name="output")
                    logger.warning("No results to save")
                    return

                # 키워드별 시트 생성
                for kw, items in results_by_keyword.items():
                    rows = [_article_to_row(item, fallback_keyword=kw) for item in items]
                    df = pd.DataFrame(rows)
                    
                    # 시트명 정리
                    sheet_name = ExcelOutputHandler._clean_sheet_name(kw or "output")
                    
                    if df.empty:
                        df = pd.DataFrame(columns=list(_article_to_row({}, kw).keys()))
                        
                    df.to_excel(writer, index=False, sheet_name=sheet_name)
                    logger.info(f"Saved {len(df)} articles for keyword '{kw}' to sheet '{sheet_name}'")
                    
            logger.info(f"Excel file saved: {path}")
            
        except Exception as e:
            logger.error(f"Failed to save Excel file: {e}")
            raise

    @staticmethod
    def save_results_to_sheet(
        results_by_keyword: Dict[str, List[Any]],
        file_path: str,
        sheet_name: str = "output",
        replace: bool = True,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
    ) -> None:
        """
        모든 키워드 결과를 하나의 시트에 저장
        """
        # 모든 결과를 하나의 리스트로 합치기
        all_rows = []
        for kw, items in results_by_keyword.items():
            for item in items:
                all_rows.append(_article_to_row(item, fallback_keyword=kw))
        
        df = pd.DataFrame(all_rows)
        if df.empty:
            df = pd.DataFrame(columns=list(_article_to_row({}, "").keys()))

        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        try:
            if path.exists():
                # 기존 파일 업데이트
                book = load_workbook(path)
                
                # 시트 교체
                if sheet_name in book.sheetnames and replace:
                    book.remove(book[sheet_name])
                
                with pd.ExcelWriter(path, engine="openpyxl", mode="a") as writer:
                    writer.book = book
                    df.to_excel(writer, index=False, sheet_name=sheet_name)

                    # meta 시트 업데이트
                    meta_data = [
                        {"key": "updated_at", "value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
                        {"key": "date_from", "value": date_from or ""},
                        {"key": "date_to", "value": date_to or ""},
                        {"key": "total_rows", "value": len(df)},
                        {"key": "sheet_name", "value": sheet_name},
                    ]
                    meta_df = pd.DataFrame(meta_data)
                    
                    if "meta" in writer.book.sheetnames and replace:
                        writer.book.remove(writer.book["meta"])
                    meta_df.to_excel(writer, index=False, sheet_name="meta")
                    
            else:
                # 새 파일 생성
                with pd.ExcelWriter(path, engine="openpyxl") as writer:
                    df.to_excel(writer, index=False, sheet_name=sheet_name)
                    
                    meta_data = [
                        {"key": "created_at", "value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
                        {"key": "date_from", "value": date_from or ""},
                        {"key": "date_to", "value": date_to or ""},
                        {"key": "total_rows", "value": len(df)},
                        {"key": "sheet_name", "value": sheet_name},
                    ]
                    meta_df = pd.DataFrame(meta_data)
                    meta_df.to_excel(writer, index=False, sheet_name="meta")
                    
            logger.info(f"Saved {len(df)} total articles to sheet '{sheet_name}' in {path}")
            
        except Exception as e:
            logger.error(f"Failed to save to sheet: {e}")
            raise

    @staticmethod
    def _clean_sheet_name(name: str, max_length: int = 31) -> str:
        """엑셀 시트명 정리"""
        # 엑셀에서 허용하지 않는 문자 제거
        invalid_chars = r'[\/\\\?\*\[\]\:]'
        safe_name = re.sub(invalid_chars, "_", str(name)).strip()
        
        # 길이 제한
        if len(safe_name) > max_length:
            safe_name = safe_name[:max_length - 3] + "..."
            
        return safe_name or "Sheet"

    @staticmethod
    def apply_styles(file_path: str, sheet_name: str = None):
        """엑셀 스타일 적용"""
        try:
            book = load_workbook(file_path)
            sheets = [book[sheet_name]] if sheet_name else book.worksheets
            
            for ws in sheets:
                # 헤더 스타일
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # 컬럼 너비 자동 조정
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            book.save(file_path)
            logger.info(f"Styles applied to {file_path}")
            
        except Exception as e:
            logger.warning(f"Could not apply styles: {e}")


# 편의 함수
def save_keyword_results(
    articles_by_keyword: Dict[str, List[Any]],
    output_dir: str = "outputs",
    filename_prefix: str = "keyword_results",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    apply_style: bool = True
) -> str:
    """
    크롤링 결과를 타임스탬프가 포함된 엑셀 파일로 저장
    """
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = output_path / f"{filename_prefix}_{timestamp}.xlsx"

    handler = ExcelOutputHandler()
    handler.save_results(
        results_by_keyword=articles_by_keyword,
        file_path=str(filepath),
        date_from=date_from,
        date_to=date_to,
    )
    
    if apply_style:
        handler.apply_styles(str(filepath))

    return str(filepath)